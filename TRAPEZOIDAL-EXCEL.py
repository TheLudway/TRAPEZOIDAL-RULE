from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


wb = load_workbook('SPEED.xlsx')
ws = wb.active
Resultado_suma = 0

#RECORRIDO DE LAS CELDAS PARA LA MULTIPLICACIÓN POR (2) EN LA SUMA
for row in range(3, 40981):
    for col in [2]:
        char = get_column_letter(col)
        Resultado_suma += (ws[char + str(row)].value) * 2

#AGREGAR LOS PRIMEROS VALORES A LA SUMA INTERIOR
Resultado_suma += ws["B2"].value + ws["B40981"].value

#OBTENER LOS LÍMITES DE INTEGRACIÓN Y CÀLCULAR DELTA DE X
Delta_x = (ws["A40981"].value - ws["A2"].value) / 40980

Resultado_final = (Resultado_suma * Delta_x) / 2
print(Resultado_final)

